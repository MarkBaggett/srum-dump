from openpyxl.cell import WriteOnlyCell, Cell
from openpyxl.styles import Font
from Registry import Registry
from datetime import datetime,timedelta
import pyesedb
import sys
import struct
import re
import openpyxl
import argparse
import warnings
import hashlib
import random
import os
import codecs
import itertools
import pathlib
import uuid
import webbrowser
import PySimpleGUI as sg
import tempfile
import urllib.request
import subprocess
import ctypes
import time


def BinarySIDtoStringSID(sid_str):
    #Original form Source: https://github.com/google/grr/blob/master/grr/parsers/wmi_parser.py
    """Converts a binary SID to its string representation.
     https://msdn.microsoft.com/en-us/library/windows/desktop/aa379597.aspx
    The byte representation of an SID is as follows:
      Offset  Length  Description
      00      01      revision
      01      01      sub-authority count
      02      06      authority (big endian)
      08      04      subauthority #1 (little endian)
      0b      04      subauthority #2 (little endian)
      ...
    Args:
      sid: A byte array.
    Returns:
      SID in string form.
    Raises:
      ValueError: If the binary SID is malformed.
    """
    if not sid_str:
        return ""
    sid = codecs.decode(sid_str,"hex")
    str_sid_components = [sid[0]]
    # Now decode the 48-byte portion
    if len(sid) >= 8:
        subauthority_count = sid[1]
        identifier_authority = struct.unpack(">H", sid[2:4])[0]
        identifier_authority <<= 32
        identifier_authority |= struct.unpack(">L", sid[4:8])[0]
        str_sid_components.append(identifier_authority)
        start = 8
        for i in range(subauthority_count):
            authority = sid[start:start + 4]
            if not authority:
                break
            if len(authority) < 4:
                raise ValueError("In binary SID '%s', component %d has been truncated. "
                         "Expected 4 bytes, found %d: (%s)",
                         ",".join([str(ord(c)) for c in sid]), i,
                         len(authority), authority)
            str_sid_components.append(struct.unpack("<L", authority)[0])
            start += 4
            sid_str = "S-%s" % ("-".join([str(x) for x in str_sid_components]))
    sid_name = template_lookups.get("Known SIDS",{}).get(sid_str,'unknown')
    return "{} ({})".format(sid_str,sid_name)

def blob_to_string(binblob):
    """Takes in a binary blob hex characters and does its best to convert it to a readable string.
       Works great for UTF-16 LE, UTF-16 BE, ASCII like data. Otherwise return it as hex.
    """
    try:
        chrblob = codecs.decode(binblob,"hex")
    except:
        chrblob = binblob
    try:
        if re.match(b'^(?:[^\x00]\x00)+\x00\x00$', chrblob):
            binblob = chrblob.decode("utf-16-le").strip("\x00")
        elif re.match(b'^(?:\x00[^\x00])+\x00\x00$', chrblob):
            binblob = chrblob.decode("utf-16-be").strip("\x00")
        else:
            binblob = chrblob.decode("latin1").strip("\x00")
    except:
        binblob = "" if not binblob else codecs.decode(binblob,"latin-1")
    return binblob

def ole_timestamp(binblob):
    """converts a hex encoded OLE time stamp to a time string"""
    try:
        td,ts = str(struct.unpack("<d",binblob)[0]).split(".")
        dt = datetime(1899,12,30,0,0,0) + timedelta(days=int(td),seconds=86400 * float("0.{}".format(ts)))
    except:
        dt = "This field is incorrectly identified as an OLE timestamp in the template."
    return dt
 
def file_timestamp(binblob):
    """converts a hex encoded windows file time stamp to a time string"""
    try:
        dt = datetime(1601,1,1,0,0,0) + timedelta(microseconds=binblob/10)
    except:
        dt = "This field is incorrectly identified as a file timestamp in the template"
    return dt

def load_registry_sids(reg_file):
    """Given Software hive find SID usernames"""
    sids = {}
    profile_key = r"Microsoft\Windows NT\CurrentVersion\ProfileList"
    tgt_value = "ProfileImagePath"
    try:
        reg_handle = Registry.Registry(reg_file)
        key_handle = reg_handle.open(profile_key)
        for eachsid in key_handle.subkeys():
            sids_path = eachsid.value(tgt_value).value()
            sids[eachsid.name()] = sids_path.split("\\")[-1]
    except:
        return {}
    return sids

def load_interfaces(reg_file):
    """Loads the names of the wireless networks from the software registry hive"""
    try:
        reg_handle = Registry.Registry(reg_file)
    except Exception as e:
        print(r"I could not open the specified SOFTWARE registry key. It is usually located in \Windows\system32\config.  This is an optional value.  If you cant find it just dont provide one.")
        print(("WARNING : ", str(e)))
        return {}
    try:
        int_keys = reg_handle.open('Microsoft\\WlanSvc\\Interfaces')
    except Exception as e:
        print("There doesn't appear to be any wireless interfaces in this registry file.")
        print(("WARNING : ", str(e)))
        return {}
    profile_lookup = {}
    for eachinterface in int_keys.subkeys():
        if len(eachinterface.subkeys())==0:
            continue
        for eachprofile in eachinterface.subkey("Profiles").subkeys():
            profileid = [x.value() for x in list(eachprofile.values()) if x.name()=="ProfileIndex"][0]
            metadata = list(eachprofile.subkey("MetaData").values())
            for eachvalue in metadata:
                if eachvalue.name() in ["Channel Hints", "Band Channel Hints"]:
                    channelhintraw = eachvalue.value()
                    hintlength = struct.unpack("I", channelhintraw[0:4])[0]
                    name = channelhintraw[4:hintlength+4] 
                    profile_lookup[str(profileid)] = name.decode(encoding="latin1")
    return profile_lookup

def load_srumid_lookups(database):
    """loads the SRUMID numbers from the SRUM database"""
    id_lookup = {}
    #Note columns  0 = Type, 1 = Index, 2 = Value
    lookup_table = database.get_table_by_name('SruDbIdMapTable')
    column_lookup = dict([(x.name,index) for index,x in enumerate(lookup_table.columns)])
    num_lookups = ese_table_record_count(lookup_table)
    if not num_lookups:
        print(f"\nUnexpectedly. The number of records in the lookup table is zero.")
        return ""
    for rec_entry_num in range(lookup_table.number_of_records):
        bin_blob = smart_retrieve(lookup_table,rec_entry_num, column_lookup['IdBlob'])
        if smart_retrieve(lookup_table,rec_entry_num, column_lookup['IdType'])==3:
            bin_blob = BinarySIDtoStringSID(bin_blob)
        elif not bin_blob == "Empty":
            bin_blob = blob_to_string(bin_blob)
        id_lookup[smart_retrieve(lookup_table,rec_entry_num, column_lookup['IdIndex'])] = bin_blob
    return id_lookup

def load_template_lookups(template_workbook):
    """Load any tabs named lookup-xyz form the template file for lookups of columns with the same format type"""
    template_lookups = {}
    for each_sheet in template_workbook.get_sheet_names():
        if each_sheet.lower().startswith("lookup-"):
            lookupname = each_sheet.split("-")[1]
            template_sheet = template_workbook.get_sheet_by_name(each_sheet)
            lookup_table = {}
            for eachrow in range(1,template_sheet.max_row+1):
                value = template_sheet.cell(row = eachrow, column = 1).value
                description = template_sheet.cell(row = eachrow, column = 2).value
                lookup_table[value] = description
            template_lookups[lookupname] = lookup_table
    return template_lookups
    
def load_template_tables(template_workbook):
    """Load template tabs that define the field names and formats for tables found in SRUM"""
    template_tables = {}    
    sheets = template_workbook.get_sheet_names()
    for each_sheet in sheets:
        #open the first sheet in the template
        template_sheet = template_workbook.get_sheet_by_name(each_sheet)
        #retieve the name of the ESE table to populate the sheet with from A1
        ese_template_table = template_sheet.cell(row=1,column=1).value
        #retrieve the names of the ESE table columns and cell styles from row 2 and format commands from row 3 
        template_field = {}
        #Read the first Row B & C in the template into lists so we know what data we are to extract
        for eachcolumn in range(1,template_sheet.max_column+1):
            field_name = template_sheet.cell(row = 2, column = eachcolumn).value
            if field_name == None:
                break
            template_style = template_sheet.cell(row = 4, column = eachcolumn).style
            template_format = template_sheet.cell(row = 3, column = eachcolumn).value
            template_value = template_sheet.cell(row = 4, column = eachcolumn ).value
            if not template_value:
                template_value= field_name
            template_field[field_name] = (template_style,template_format,template_value)
        template_tables[ese_template_table] = (each_sheet, template_field)
    return template_tables    


def smart_retrieve(ese_table, ese_record_num, column_number):
    """Given a row and column will determine the format and retrieve a value from the ESE table"""
    rec = ese_table.get_record(ese_record_num)
    col_type = rec.get_column_type(column_number)
    col_data = rec.get_value_data(column_number)
    #print "rec:%s  col:%s type:%s %s" % (ese_record_num, column_number, col_type, ese_column_types[col_type])
    if col_type == pyesedb.column_types.BINARY_DATA:
        col_data = "" if not col_data else codecs.encode(col_data,"HEX")
    elif col_type == pyesedb.column_types.BOOLEAN:
        col_data = struct.unpack('?',col_data)[0]
    elif col_type == pyesedb.column_types.CURRENCY:
        pass
    elif col_type == pyesedb.column_types.DATE_TIME:
        col_data = ole_timestamp(col_data)
    elif col_type == pyesedb.column_types.DOUBLE_64BIT:
        col_data = 0 if not col_data else struct.unpack('d',col_data)[0]
    elif col_type == pyesedb.column_types.FLOAT_32BIT:
        col_data = 0.0 if not col_data else struct.unpack('f',col_data)[0]
    elif col_type == pyesedb.column_types.GUID:
        col_data = 0 if not col_data else str(uuid.UUID(bytes = col_data))
    elif col_type == pyesedb.column_types.INTEGER_16BIT_SIGNED:
        col_data = 0 if not col_data else struct.unpack('h',col_data)[0]
    elif col_type == pyesedb.column_types.INTEGER_16BIT_UNSIGNED:
        col_data = 0 if not col_data else struct.unpack('H',col_data)[0]
    elif col_type == pyesedb.column_types.INTEGER_32BIT_SIGNED:
        col_data =  0 if not col_data else struct.unpack('i',col_data)[0]
    elif col_type == pyesedb.column_types.INTEGER_32BIT_UNSIGNED:
        col_data = 0 if not col_data else struct.unpack('I',col_data)[0]
    elif col_type == pyesedb.column_types.INTEGER_64BIT_SIGNED:
        col_data = 0 if not col_data else struct.unpack('q',col_data)[0]
    elif col_type == pyesedb.column_types.INTEGER_8BIT_UNSIGNED:
        col_data = 0 if not col_data else struct.unpack('B',col_data)[0]
    elif col_type == pyesedb.column_types.LARGE_BINARY_DATA:
        col_data = "" if not col_data else codecs.encode(col_data,"HEX")
    elif col_type == pyesedb.column_types.LARGE_TEXT:
        col_data = blob_to_string(col_data)
    elif col_type == pyesedb.column_types.NULL:
        pass
    elif col_type == pyesedb.column_types.SUPER_LARGE_VALUE:
        col_data = "" if not col_data else codecs.encode(col_data,"HEX")
    elif col_type == pyesedb.column_types.TEXT:
        col_data = blob_to_string(col_data)  
    else:
        col_data = blob_to_string(col_data)    
    if col_data==None:
        col_data = "Empty"
    return col_data

def format_output(val, eachformat, eachstyle, xls_sheet):
    """Returns a excel cell with the data formated as specified in the template table"""
    new_cell = WriteOnlyCell(xls_sheet, value = "init")
    new_cell.style = eachstyle
    if val==None:
        val="None"
    elif eachformat in [None, "OLE"]:
        pass
    elif eachformat.startswith("OLE:"):
        val = val.strftime(eachformat[4:])
    elif eachformat=="FILE":
        val = file_timestamp(val)
        new_cell.number_format = 'YYYY MMM DD'
    elif eachformat.startswith("FILE:"):
        val = file_timestamp(val)
        val = val.strftime(eachformat[5:])
    elif eachformat.lower().startswith("lookup-"):
        lookup_name = eachformat.split("-")[1]
        if lookup_name in template_lookups:
            lookup_table = template_lookups.get(lookup_name,{})
            val = lookup_table.get(val,val)
    elif eachformat.lower() == "lookup_id":
        val = id_table.get(val, "No match in srum lookup table for %s" % (val))
    elif eachformat.lower() == "lookup_luid":
        inttype = struct.unpack(">H6B", codecs.decode(format(val,'016x'),'hex'))[0]
        val = template_lookups.get("LUID Interfaces",{}).get(inttype,"")
    elif eachformat.lower() == "seconds":
        val = val/86400.0
        new_cell.number_format = 'dd hh:mm:ss'
    elif eachformat.lower() == "md5":
        val = hashlib.md5(str(val)).hexdigest()
    elif eachformat.lower() == "sha1":
        val = hashlib.sha1(str(val)).hexdigest()
    elif eachformat.lower() == "sha256":
        val = hashlib.sha256(str(val)).hexdigest()
    elif eachformat.lower() == "base16":
        if type(val)==int:
            val = hex(val)
        else:
            val = format(val,"08x")
    elif eachformat.lower() == "base2":
        if type(val)==int:
            val = format(val,"032b")
        else:
            try:
                val = int(str(val),2)
            except :
                val = val
    elif eachformat.lower() == "interface_id" and options.reghive:
        val = interface_table.get(str(val),"")
    elif eachformat.lower() == "interface_id" and not options.reghive:
        val = val
    else:
        val = val
    try:
        new_cell.value = val
    except:
        new_cell.value = re.sub(r'[\000-\010]|[\013-\014]|[\016-\037]|[\x00-\x1f\x7f-\x9f]|[\uffff]',"",val)
    return new_cell

def ese_table_guid_to_name(ese_table):
    if ese_table.name in template_tables:
        tname,tfields = template_tables.get(ese_table.name)
    else:
        tname = ese_table.get_name()
    return tname

def ese_table_get_record(ese_table, row_num):
    retry = 5
    if row_num >= ese_table_record_count(ese_table):
        return None
    while retry:
        try:
            ese_row = ese_table.get_record(row_num)
        except Exception as e:
            retry -= 1
            time.sleep(0.1)
            error = e
        else:
            break
    else:
        tname = ese_table_guid_to_name(ese_table)
        print("Skipping corrupt row {0} in the {1} table. Because {2}".format(row_num, tname, str(error)))
        ese_row = None
    return ese_row

def ese_table_record_count(ese_table):
    retry = 5
    while retry:
        try:
            total_recs = ese_table.get_number_of_records()
        except:
            retry -= 1
            time.sleep(0.1)
        else:
            break
    else:
        tname = ese_table_guid_to_name(x)
        print(f"Table {tname} has an invalid number of records. {str(total_recs)}")
        total_recs = 0
    return total_recs


def process_srum(ese_db, target_wb ):
    """Process all the tables and columns in the ESE database"""
    total_recs = 0
    for each_table in  ese_db.tables:
        if each_table.name in skip_tables:
            continue
        total_recs += ese_table_record_count(each_table)

    if not options.quiet:
        print("Processing {} records across {} tables".format(total_recs,ese_db.number_of_tables-len(skip_tables)))
    for table_num in range(ese_db.number_of_tables):
        ese_table = ese_db.get_table(table_num)
        if ese_table.name in skip_tables:
            continue

        tname = ese_table_guid_to_name(ese_table)
        num_recs = ese_table_record_count(ese_table)
        if not num_recs:
            print(f"\nSkipping table with zero of records. {tname}")
            continue

        if not options.quiet:
            print("\nNow dumping table {} containing {} rows".format(tname, num_recs or "Unknown"))
            print("While you wait, did you know ...\n {} \n".format(next(ads)))

        xls_sheet = target_wb.create_sheet(title=tname)

        column_names = [x.name for x in ese_table.columns]
        column_widths = [len(x.name)+2 for x in ese_table.columns]

        header_row = [x.name for x in ese_table.columns]
        if ese_table.name in template_tables:
            tname,tfields = template_tables.get(ese_table.name)
            header_row = []
            for i, eachcol in enumerate(ese_table.columns):
                if eachcol.name in tfields:
                    cell_style, _, cell_value = tfields.get(eachcol.name)
                    new_cell = WriteOnlyCell(xls_sheet, value=cell_value)
                    new_cell.style = cell_style
                    header_row.append( new_cell )
                    column_widths[i] = len(str(cell_value)) + 2
                else:
                    header_row.append(WriteOnlyCell(xls_sheet, value=eachcol.name))
        xls_sheet.append(header_row)
    
        for row_num in range(num_recs):
            
            ese_row = ese_table_get_record(ese_table,row_num)
            if ese_row == None:
                continue

            if not options.quiet and row_num % 500 == 0:
                 print("\r|{0:-<50}| {1:3.2f}%".format("X"*( 50 * row_num//num_recs), 100*row_num/num_recs ),end="")

                #The row is retrieved now use the template to figure out which ones you want and format them
            xls_row = []
            for col_num in range(ese_table.number_of_columns):
                val = smart_retrieve(ese_table,row_num, col_num)
                if val=="Error":
                    val = "WARNING: Invalid Column Name {}".format(column_names[col_num])
                elif val==None:
                    val="None"  
                elif ese_table.name in template_tables:
                    tname,tfields = template_tables.get(ese_table.name) 
                    if column_names[col_num] in tfields:
                        cstyle, cformat, _ = tfields.get(column_names[col_num])
                        val = format_output(val, cformat, cstyle,xls_sheet)
                val_len = len(str(val.value)) if isinstance(val, Cell) else len(str(val))
                if column_widths[col_num] < val_len:
                    column_widths[col_num] = val_len
                #print dir(new_cell.style.font)
                xls_row.append(val)
            xls_sheet.append(xls_row)
        # Adjust column widths now
        if ese_table.name in template_tables:
            tname,tfields = template_tables.get(ese_table.name)
            counter = 1
            for i in range(ese_table.number_of_columns):
                if column_names[i] in tfields:
                    xls_sheet.column_dimensions[openpyxl.utils.get_column_letter(counter)].width = min(column_widths[i], 80)
                counter += 1
        else:
            for i in range(ese_table.number_of_columns):
                xls_sheet.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = min(column_widths[i], 80)
        # Add AutoFilter
        xls_sheet.auto_filter.ref = xls_sheet.dimensions
        # Bold first row
        for x in range(0, len(header_row)):
            xls_sheet[f'{openpyxl.utils.get_column_letter(x+1)}1'].font = Font(bold=True)
        # Freeze first row
        xls_sheet.freeze_panes = "A2"
        if not options.quiet:
            print("\r|XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX| 100.00% FINISHED")

def show_live_system_warning():
    """Warn the user when they try to analyze the srum on their own live system."""
    layout = [
          [sg.Text("It appears your trying to open SRUDB.DAT from a live system.")],
          [sg.Text("Copying or reading that file while it is locked is unlikely to succeed.")],
          [sg.Text("First, use a tool such as FGET that can copy files that are in use.")], 
          [sg.Text(r"Try: 'fget -extract c:\windows\system32\sru\srudb.dat <a destination path>'")],
          [sg.Button("Close"), sg.Button("Download FGET") ]
         ]
    if ctypes.windll.shell32.IsUserAnAdmin() == 1:
        layout[-1].append(sg.Button("Auto Extract"))
    pop_window = sg.Window("WARNING", layout, no_titlebar=True, keep_on_top=True, border_depth=5)
    return_value = None
    while True:
        event,_  = pop_window.Read()
        if event in (None,"Close"):
            break
        if event == "Download FGET":
            webbrowser.open("https://github.com/MarkBaggett/srum-dump/blob/master/FGET.exe")
        if event == "Auto Extract":
            return_value = extract_live_file()
            break
    pop_window.Close()
    return return_value

def extract_live_file():
    try:
        tmp_dir = tempfile.mkdtemp()
        #fget_file = tempfile.NamedTemporaryFile(mode="w+b", suffix=".exe",delete=False)
        fget_file = pathlib.Path(tmp_dir) / "fget.exe"
        #registry_file = tempfile.NamedTemporaryFile(mode="w+b", suffix = ".reg", delete=False)
        registry_file = pathlib.Path(tmp_dir) / "SOFTWARE"
        #extracted_srum = tempfile.NamedTemporaryFile(mode="w+b", suffix = ".dat", delete=False
        extracted_srum = pathlib.Path(tmp_dir) / "srudb.dat"
        esentutl_path = pathlib.Path(os.environ.get("COMSPEC")).parent / "esentutl.exe"
        if esentutl_path.exists():
            print("Extracting srum with esentutl.exe")
            cmdline = r"{} /y c:\\windows\\system32\\sru\\srudb.dat /vss /d {}".format(str(esentutl_path), str(extracted_srum))
            print(cmdline)
            phandle = subprocess.Popen(cmdline, shell=True,stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            out1,_ = phandle.communicate()
            print("Extracting Registry with esentutl.exe")
            cmdline = r"{} /y c:\\windows\\system32\\config\\SOFTWARE /vss /d {}".format(str(esentutl_path), str(registry_file))
            print(cmdline)
            phandle = subprocess.Popen(cmdline, shell=True,stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            out2,_ = phandle.communicate()
        else:
            print("Downloading fget.exe to {}".format(str(fget_file)))
            fget_binary = urllib.request.urlopen('https://github.com/MarkBaggett/srum-dump/raw/master/FGET.exe').read()
            fget_file.write_bytes(fget_binary)
            print("Extracting srum with fget.exe")
            cmdline = r"{} -extract c:\\windows\\system32\\sru\srudb.dat {}".format(str(fget_file), str(extracted_srum))
            print(cmdline)
            phandle = subprocess.Popen(cmdline, shell=True,stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            out1,_ = phandle.communicate()
            cmdline = r"{} -extract c:\\windows\\system32\\config\SOFTWARE {}".format(str(fget_file), str(registry_file))
            print(cmdline)
            phandle = subprocess.Popen(cmdline, shell=True,stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            out2,_ = phandle.communicate()
            fget_file.unlink()
    except Exception as e:
        print("Unable to automatically extract srum. {}\n{}\n{}".format(str(e), out1.decode(), out2.decode()))
        return None
    if (b"returned error" in out1+out2) or (b"Init failed" in out1+out2):
        print("ERROR\n SRUM Extraction: {}\n Registry Extraction {}".format(out1.decode(),out2.decode()))
    elif b"success" in out1.lower() and b"success" in out2.lower():
        return str(extracted_srum), str(registry_file)
    else:
        print("Unable to determine success or failure.", out1.decode(),"\n",out2.decode())
    return None
 
parser = argparse.ArgumentParser(description="Given an SRUM database it will create an XLS spreadsheet with analysis of the data in the database.")
parser.add_argument("--SRUM_INFILE","-i", help ="Specify the ESE (.dat) file to analyze. Provide a valid path to the file.")
parser.add_argument("--XLSX_OUTFILE", "-o", default="SRUM_DUMP_OUTPUT.xlsx", help="Full path to the XLS file that will be created.")
parser.add_argument("--XLSX_TEMPLATE" ,"-t", help = "The Excel Template that specifies what data to extract from the srum database. You can create template_tables with ese_template.py.")
parser.add_argument("--REG_HIVE", "-r", dest="reghive", help = "If SOFTWARE registry hive is provided then the names of the network profiles will be resolved.")
parser.add_argument("--quiet", "-q", help = "Supress unneeded output messages.",action="store_true")
options = parser.parse_args()

ads = itertools.cycle(["Did you know SANS Automating Infosec with Python SEC573 teaches you to develop Forensics and Incident Response tools?",
       "To learn how SRUM and other artifacts can enhance your forensics investigations check out SANS Windows Forensic Analysis FOR500.",
       "Yogesh Khatri's paper at https://github.com/ydkhatri/Presentations/blob/master/SRUM%20Forensics-SANS.DFIR.summit.2015.pdf was essential in the creation of this tool.",
       "By modifying the template file you have control of what ends up in the analyzed results.  Try creating an alternate template and passing it with the --XLSX_TEMPLATE option.",
       "TIP: When using a SOFTWARE registry file you can add your own SIDS to the 'lookup-Known SIDS' tab!",
       "This program was written by Twitter:@markbaggett and @donaldjwilliam5 because @ovie said so.",
       "SRUM-DUMP 2.0 will attempt to dump any ESE database! If no template defines a table it will do its best to guess."
       ])

if not options.SRUM_INFILE:
    srum_path = ""
    if os.path.exists("SRUDB.DAT"):
        srum_path = os.path.join(os.getcwd(),"SRUDB.DAT")
    temp_path = pathlib.Path.cwd() / "SRUM_TEMPLATE2.XLSX"
    if temp_path.exists():
        temp_path = str(temp_path)
    else:
        temp_path = ""
    reg_path = ""
    if os.path.exists("SOFTWARE"):
        reg_path = os.path.join(os.getcwd(),"SOFTWARE")

    sg.ChangeLookAndFeel('DarkRed2')
    layout = [[sg.Text('REQUIRED: Path to SRUDB.DAT')],
    [sg.Input(srum_path,key="_SRUMPATH_", enable_events=True), sg.FileBrowse(target="_SRUMPATH_")], 
    [sg.Text('REQUIRED: Output folder for SRUM_DUMP_OUTPUT.xlsx')],
    [sg.Input(os.getcwd(),key='_OUTDIR_'), sg.FolderBrowse(target='_OUTDIR_')],
    [sg.Text('REQUIRED: Path to SRUM_DUMP Template')],
    [sg.Input(temp_path,key="_TEMPATH_"), sg.FileBrowse(target="_TEMPATH_")],
    [sg.Text('RECOMMENDED: Path to registry SOFTWARE hive')],
    [sg.Input(key="_REGPATH_"), sg.FileBrowse(target="_REGPATH_")],
    [sg.Text("Click here for support via Twitter @MarkBaggett",enable_events=True, key="_SUPPORT_", text_color="Blue")],
    [sg.OK(), sg.Cancel()]] 
    
    # Create the Window
    window = sg.Window('SRUM_DUMP 2.6', layout)

    while True:             
        event, values = window.Read()
        if event is None:
            break
        if event == "_SUPPORT_":
            webbrowser.open("https://twitter.com/MarkBaggett")
        if event == 'Cancel':
            sys.exit(0)
        if event == "_SRUMPATH_":
            if str(pathlib.Path(values.get("_SRUMPATH_"))).lower() == "c:\\windows\\system32\\sru\\srudb.dat":
                result = show_live_system_warning() 
                if result:
                    window.Element("_SRUMPATH_").Update(result[0])
                    window.Element("_REGPATH_").Update(result[1])
                continue
        if event == 'OK':
            tmp_path = pathlib.Path(values.get("_SRUMPATH_"))
            if not tmp_path.exists() or not tmp_path.is_file():
                sg.PopupOK("SRUM DATABASE NOT FOUND.")
                continue
            if not os.path.exists(pathlib.Path(values.get("_OUTDIR_"))):
                sg.PopupOK("OUTPUT DIR NOT FOUND.")
                continue
            tmp_path = pathlib.Path(values.get("_TEMPATH_"))            
            if not tmp_path.exists() or not tmp_path.is_file():
                sg.PopupOK("SRUM TEMPLATE NOT FOUND.")
                continue
            tmp_path = pathlib.Path(values.get("_REGPATH_"))
            if values.get("_REGPATH_") and not tmp_path.exists() and not tmp_path.is_file():
                sg.PopupOK("REGISTRY File not found. (Leave field empty for None.)")
                continue
            break

    window.Close()
    options.SRUM_INFILE = str(pathlib.Path(values.get("_SRUMPATH_")))
    options.XLSX_OUTFILE = str(pathlib.Path(values.get("_OUTDIR_")) / "SRUM_DUMP_OUTPUT.xlsx")
    options.XLSX_TEMPLATE = str(pathlib.Path(values.get("_TEMPATH_")))
    options.reghive = str(pathlib.Path(values.get("_REGPATH_")))
    if options.reghive == ".":
        options.reghive = ""
else:
    if not options.XLSX_TEMPLATE:
        options.XLSX_TEMPLATE = "SRUM_TEMPLATE2.xlsx"
    if not options.XLSX_OUTFILE:
        options.XLSX_OUTFILE = "SRUM_DUMP_OUTPUT.xlsx"
    if not os.path.exists(options.SRUM_INFILE):
        print("ESE File Not found: "+options.SRUM_INFILE)
        sys.exit(1)
    if not os.path.exists(options.XLSX_TEMPLATE):
        print("Template File Not found: "+options.XLSX_TEMPLATE)
        sys.exit(1)
    if options.reghive and not os.path.exists(options.reghive):
        print("Registry File Not found: "+options.reghive)
        sys.exit(1)

regsids = {}
if options.reghive:
    interface_table = load_interfaces(options.reghive)
    regsids = load_registry_sids(options.reghive)

try:
    warnings.simplefilter("ignore")
    ese_db = pyesedb.file()
    ese_db.open(options.SRUM_INFILE)
    #ese_db = ese.ESENT_DB(options.SRUM_INFILE)
except Exception as e:
    print("I could not open the specified SRUM file. Check your path and file name.")
    print("Error : ", str(e))
    sys.exit(1) 

try:
    template_wb = openpyxl.load_workbook(filename=options.XLSX_TEMPLATE)
except Exception as e:
    print("I could not open the specified template file %s. Check your path and file name." % (options.XLSX_TEMPLATE))
    print("Error : ", str(e))
    sys.exit(1)

skip_tables = ['MSysObjects', 'MSysObjectsShadow', 'MSysObjids', 'MSysLocales','SruDbIdMapTable']
template_tables = load_template_tables(template_wb)
template_lookups = load_template_lookups(template_wb)
if regsids:
    template_lookups.get("Known SIDS",{}).update(regsids)
    #print("REGSIDS!!!")
    #print(template_lookups.get("Known SIDS"))
id_table = load_srumid_lookups(ese_db)

target_wb = openpyxl.Workbook()
process_srum(ese_db, target_wb)

firstsheet=target_wb.get_sheet_by_name("Sheet")
target_wb.remove_sheet(firstsheet)
print("Writing output file to disk.")
try:
    target_wb.save(options.XLSX_OUTFILE)
except Exception as e:
    print("I was unable to write the output file.  Do you have an old version open?  If not this is probably a path or permissions issue.")
    print("Error : ", str(e))

print("Done.")
