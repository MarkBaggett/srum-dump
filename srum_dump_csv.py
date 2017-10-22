#
# SRUM DUMP that uses different ESE module and only dumps CSVs.  Some large
# SRUM files crashed the XLSX module.
#
"""SRUM DUMP CSV Only.  Mark Baggett."""

import pyesedb
from datetime import datetime, timedelta
import os
import sys
import struct
import openpyxl
import argparse
import warnings
import hashlib
from copy import copy
import csv


def lookup_sid(sid):
    """Given a SID looks up the common account name."""
    # Returns a common sid https://support.microsoft.com/en-us/kb/243330
    known_sids = {'S-1-5-32-545': ' Users', 'S-1-5-32-544': ' Administrators', 'S-1-5-32-547': ' Power Users', 'S-1-5-32-546': ' Guests', 'S-1-5-32-569': ' BUILTIN\\Cryptographic Operators', 'S-1-16-16384': ' System Mandatory Level ', 'S-1-5-32-551': ' Backup Operators', 'S-1-16-8192': ' Medium Mandatory Level ', 'S-1-5-80': ' NT Service ', 'S-1-5-32-548': ' Account Operators', 'S-1-5-32-561': ' BUILTIN\\Terminal Server License Servers', 'S-1-5-64-14': ' SChannel Authentication ', 'S-1-5-32-562': ' BUILTIN\\Distributed COM Users', 'S-1-5-64-21': ' Digest Authentication ', 'S-1-5-19': ' NT Authority', 'S-1-3-0': ' Creator Owner', 'S-1-5-80-0': ' All Services ', 'S-1-5-20': ' NT Authority', 'S-1-5-18': ' Local System', 'S-1-5-32-552': ' Replicators', 'S-1-5-32-579': ' BUILTIN\\Access Control Assistance Operators', 'S-1-16-4096': ' Low Mandatory Level ', 'S-1-16-12288': ' High Mandatory Level ', 'S-1-2-0': ' Local ', 'S-1-16-0': ' Untrusted Mandatory Level ', 'S-1-5-3': ' Batch', 'S-1-5-2': ' Network', 'S-1-5-1': ' Dialup', 'S-1-5-7': ' Anonymous', 'S-1-5-6': ' Service', 'S-1-5-4': ' Interactive', 'S-1-5-9': ' Enterprise Domain Controllers', 'S-1-5-8': ' Proxy', 'S-1-5-32-550': ' Print Operators', 'S-1-0-0': ' Nobody', 'S-1-5-32-559': ' BUILTIN\\Performance Log Users', 'S-1-5-32-578': ' BUILTIN\\Hyper-V Administrators', 'S-1-5-32-549': ' Server Operators', 'S-1-2-1': ' Console Logon ', 'S-1-3-1': ' Creator Group', 'S-1-5-32-575': ' BUILTIN\\RDS Remote Access Servers', 'S-1-3-3': ' Creator Group Server', 'S-1-3-2': ' Creator Owner Server', 'S-1-5-32-556': ' BUILTIN\\Network Configuration Operators', 'S-1-5-32-557': ' BUILTIN\\Incoming Forest Trust Builders', 'S-1-5-32-554': ' BUILTIN\\Pre-Windows 2000 Compatible Access', 'S-1-5-32-573': ' BUILTIN\\Event Log Readers ', 'S-1-5-32-576': ' BUILTIN\\RDS Endpoint Servers', 'S-1-5-83-0': ' NT VIRTUAL MACHINE\\Virtual Machines', 'S-1-16-28672': ' Secure Process Mandatory Level ', 'S-1-5-11': ' Authenticated Users', 'S-1-1-0': ' Everyone', 'S-1-5-32-555': ' BUILTIN\\Remote Desktop Users', 'S-1-16-8448': ' Medium Plus Mandatory Level ', 'S-1-5-17': ' This Organization ', 'S-1-5-32-580': ' BUILTIN\\Remote Management Users', 'S-1-5-15': ' This Organization ', 'S-1-5-14': ' Remote Interactive Logon ', 'S-1-5-13': ' Terminal Server Users', 'S-1-5-12': ' Restricted Code', 'S-1-5-32-577': ' BUILTIN\\RDS Management Servers', 'S-1-5-10': ' Principal Self', 'S-1-3': ' Creator Authority', 'S-1-2': ' Local Authority', 'S-1-1': ' World Authority', 'S-1-0': ' Null Authority', 'S-1-5-32-574': ' BUILTIN\\Certificate Service DCOM Access ', 'S-1-5': ' NT Authority', 'S-1-4': ' Non-unique Authority', 'S-1-5-32-560': ' BUILTIN\\Windows Authorization Access Group', 'S-1-16-20480': ' Protected Process Mandatory Level ', 'S-1-5-64-10': ' NTLM Authentication ', 'S-1-5-32-558': ' BUILTIN\\Performance Monitor Users'}
    if sid in known_sids:
        return "%s (%s)" % (sid, known_sids.get(sid, sid))
    return sid


def BinarySIDtoStringSID(sid):
    """ Given a SID Binary blob returns a string."""
    # Original form Source: https://github.com/google/grr/blob/master/grr/parsers/wmi_parser.py
    """Converts a binary SID to its string representation.
    https://msdn.microsoft.com/en-us/library/windows/desktop/aa379597.aspx
    The byte representation of an SID is as follows:
        Offset  Length  Description
        00      01      revision
        01      01      sub-authority count
        02      06      authority (big endian)
        08      04      subauthority #1 (little endian)
        0b      04      subauthority #2 (little endian)
        ...
      Args:
        sid: A byte array.
      Returns:
        SID in string form.
      Raises:
        ValueError: If the binary SID is malformed.
    """
    if not sid:
        return ""
    str_sid_components = [ord(sid[0])]
    # Now decode the 48-byte portion
    if len(sid) >= 8:
        subauthority_count = ord(sid[1])
        identifier_authority = struct.unpack(">H", sid[2:4])[0]
        identifier_authority <<= 32
        identifier_authority |= struct.unpack(">L", sid[4:8])[0]
        str_sid_components.append(identifier_authority)
        start = 8
        for i in range(subauthority_count):
            authority = sid[start:start + 4]
            if not authority:
                break
            if len(authority) < 4:
                raise ValueError("In binary SID '%s', component %d has been truncated. "
                    "Expected 4 bytes, found %d: (%s)",
                    ",".join([str(ord(c)) for c in sid]), i,
                    len(authority), authority)
            str_sid_components.append(struct.unpack("<L", authority)[0])
            start += 4
            sid_str = "S-%s" % ("-".join([str(x) for x in str_sid_components]))
        return lookup_sid(sid_str)

def ole_timestamp(binblob,timeformat="%Y-%m-%d %H:%M:%S"):
    #converts a hex encoded OLE time stamp to a time string
    try:
        ts = struct.unpack("<d",binblob)[0]
    except:
        print("Unable to unpack binblob {0}".format(str(binblob)))
        ts = 2
    dt = datetime(1899, 12, 30, 0, 0, 0) + timedelta(days=ts)
    return  dt.strftime(timeformat)

def file_timestamp(binblob,timeformat="%Y-%m-%d %H:%M:%S"):
    #converts a hex encoded windows file time stamp to a time string
    dt = datetime(1601, 1, 1, 0, 0, 0) + timedelta(microseconds=binblob // 10)
    return  dt.strftime(timeformat)

def ese_getnextrow(ese_table):
    reverse_column_lookup = dict([(index,x.name) for index,x in enumerate(ese_table.columns)])
    for ese_row_num in range(ese_table.number_of_records):
        current_row = {}
        for each_column in range(len(ese_table.columns)):
            current_row[reverse_column_lookup[each_column]] = smart_retrieve(ese_table, ese_row_num, each_column)
        yield(current_row)

def smart_retrieve(ese_table, ese_record_num, column_number):
    ese_column_types = {0: 'NULL', 1: 'BOOLEAN', 2: 'INTEGER_8BIT_UNSIGNED', 3: 'INTEGER_16BIT_SIGNED', 4: 'INTEGER_32BIT_SIGNED', 5: 'CURRENCY', 6: 'FLOAT_32BIT', 7: 'DOUBLE_64BIT', 8: 'DATE_TIME', 9: 'BINARY_DATA', 10: 'TEXT', 11: 'LARGE_BINARY_DATA', 12: 'LARGE_TEXT', 13: 'SUPER_LARGE_VALUE', 14: 'INETEGER_32BIT_UNSIGNED', 15: 'INTEGER_64BIT_SIGNED', 16: 'GUID', 17: 'INTEGER_16BIT_UNSIGNED'}
    rec = ese_table.get_record(ese_record_num)
    col_type = rec.get_column_type(column_number)
    col_data = rec.get_value_data(column_number)
    #print "rec:%s  col:%s type:%s %s" % (ese_record_num, column_number, col_type, ese_column_types[col_type])
    if col_type == pyesedb.column_types.BINARY_DATA:
        col_data = "" if not col_data else col_data.encode("HEX")
    elif col_type == pyesedb.column_types.BOOLEAN:
        col_data = struct.unpack('?',col_data)[0]
    elif col_type == pyesedb.column_types.CURRENCY:
        pass
    elif col_type == pyesedb.column_types.DATE_TIME:
        col_data = ole_timestamp(col_data)
    elif col_type == pyesedb.column_types.DOUBLE_64BIT:
        col_data = 0 if not col_data else struct.unpack('d',col_data)[0]
    elif col_type == pyesedb.column_types.FLOAT_32BIT:
        col_data = 0.0 if not col_data else struct.unpack('f',col_data)[0]
    elif col_type == pyesedb.column_types.GUID:
        col_data = str(uuid.UUID(col_data.encode('hex')))    
    elif col_type == pyesedb.column_types.INTEGER_16BIT_SIGNED:
        col_data = 0 if not col_data else struct.unpack('h',col_data)[0]
    elif col_type == pyesedb.column_types.INTEGER_16BIT_UNSIGNED:
        col_data = 0 if not col_data else struct.unpack('H',col_data)[0]
    elif col_type == pyesedb.column_types.INTEGER_32BIT_SIGNED:
        col_data =  0 if not col_data else struct.unpack('i',col_data)[0]
    elif col_type == pyesedb.column_types.INTEGER_32BIT_UNSIGNED:
        col_data = 0 if not col_data else struct.unpack('I',col_data)[0]
    elif col_type == pyesedb.column_types.INTEGER_64BIT_SIGNED:
        col_data = 0 if not col_data else struct.unpack('q',col_data)[0]
    elif col_type == pyesedb.column_types.INTEGER_8BIT_UNSIGNED:
        col_data = 0 if not col_data else struct.unpack('B',col_data)[0]
    elif col_type == pyesedb.column_types.LARGE_BINARY_DATA:
        col_data = "" if not col_data else col_data.encode("HEX")
    elif col_type == pyesedb.column_types.LARGE_TEXT:
        col_data = blob_to_string(col_data)
    elif col_type == pyesedb.column_types.NULL:
        pass
    elif col_type == pyesedb.column_types.SUPER_LARGE_VALUE:
        col_data = "" if not col_data else col_data.encode("HEX")
    elif col_type == pyesedb.column_types.TEXT:
        col_data = blob_to_string(col_data)       
    if col_data==None:
        col_data = "Empty"
    return col_data
 
def load_interfaces(reg_file):
    from Registry import Registry
    try:
        reg_handle = Registry.Registry(reg_file)
    except Exception as e:
        print "I could not open the specified SOFTWARE registry key. It is usually located in \Windows\system32\config.  This is an optional value.  If you cant find it just dont provide one."
        print "WARNING : ", str(e)
        return {}
    try:
        int_keys = reg_handle.open('Microsoft\\WlanSvc\\Interfaces')
    except Exception as e:
        print "There doesn't appear to be any wireless interfaces in this registry file."
        print "WARNING : ", str(e)
        return {}
    profile_lookup = {}
    for eachinterface in int_keys.subkeys():
        if len(eachinterface.subkeys())==0:
            continue
        for eachprofile in eachinterface.subkey("Profiles").subkeys():
            profileid = [x.value() for x in eachprofile.values() if x.name()=="ProfileIndex"][0]
            metadata = eachprofile.subkey("MetaData").values()
            for eachvalue in metadata:
                if eachvalue.name()=="Channel Hints":
                    channelhintraw = eachvalue.value()
                    hintlength = struct.unpack("I", channelhintraw[0:4])[0]
                    name = channelhintraw[4:hintlength+4] 
                    profile_lookup[str(profileid)] = name
    return profile_lookup

def load_lookups(database):
    id_lookup = {}
    try:
        lookup_table = database.get_table_by_name('SruDbIdMapTable')
    except Exception as e:
        print "Unable to open the ID Lookup table.  Error :", str(e)
        abort(1)
    for rec_entry in ese_getnextrow(lookup_table):
        #except Exception as e:
        #    print "Skipping a corrupt record in SruDbIdMapTable."
        #    print "Error :", str(e)
        #    continue
        if rec_entry == None:
            return id_lookup
        if rec_entry['IdType']==0:
            proc_blob = 'None' if not rec_entry['IdBlob'] else unicode(rec_entry['IdBlob'].decode("hex"),'utf-16-le').strip("\x00")
            id_lookup[rec_entry['IdIndex']] = proc_blob
        elif rec_entry['IdType']==1:
            id_lookup[rec_entry['IdIndex']] = unicode(rec_entry['IdBlob'].decode("hex"),'utf-16-le').strip("\x00")
        elif rec_entry['IdType']==2:
            id_lookup[rec_entry['IdIndex']] = unicode(rec_entry['IdBlob'].decode("hex"),'utf-16-le').strip("\x00")
        elif rec_entry['IdType']==3:
            user_blob = 'None' if not rec_entry['IdBlob'] else BinarySIDtoStringSID(rec_entry['IdBlob'].decode("hex"))
            id_lookup[rec_entry['IdIndex']] = user_blob
        else:
            print "WARNING: Unknown entry type in IdMapTable"
            #print rec_entry
    return id_lookup

def abort(error_code):
    if interactive_mode:
        raw_input("Press enter to exit")
    sys.exit(error_code)

def format_value(val, eachformat):
    "Returns a excel cell with the data formated as specified"
    if val==None:
        val="None"
    elif eachformat == None:
        pass
    elif eachformat=="FILE":
        val = file_timestamp(val)
    elif eachformat.lower() == "lookup_id":
        val = id_table.get(val, "No match in srum lookup table for %s" % (val))
    elif eachformat.lower() == "lookup_luid":
        val = lookup_luid(val)
    elif eachformat.lower() == "lookup_sid":
        val = "%s (%s)" % (val, lookup_sid(val))
    elif eachformat.lower() == "seconds":
        val = val/86400.0
    elif eachformat.lower() == "md5":
        val = hashlib.md5(str(val)).hexdigest()
    elif eachformat.lower() == "sha1":
        val = hashlib.sha1(str(val)).hexdigest()
    elif eachformat.lower() == "sha256":
        val = hashlib.sha256(str(val)).hexdigest()
    elif eachformat.lower() == "base16":
        if type(val)=="<type 'int'>":
            val = hex(val)
        else:
            val = str(val).encode("hex")
    elif eachformat.lower() == "base2":
        if type(val)==int:
            val = bin(val)
        else:
            try:
                val = int(str(val),2)
            except :
                val = val
    elif eachformat.lower() == "interface_id" and options.reghive:
        val = interface_table.get(str(val),"")
    else:
        val = val
    return val

def rotating_list(somelist):
    while True:
        for x in somelist:
           yield x
          
ads = rotating_list(["Check out SANS Automating Infosec with Python SEC573 to learn to write program like this on your own.",
       "To learn how SRUM and other artifacts can enhance your forensics investigations check out SANS Windows Forensics FOR500/408",
       "This program uses the function BinarySIDtoStringSID from the GRR code base to convert binary data into a user SID and relies heavily on the CoreSecurity Impacket ESE module. This works because of them.  Check them out!",
       "Yogesh Khatri's paper at https://files.sans.org/summit/Digital_Forensics_and_Incident_Response_Summit_2015/PDFs/Windows8SRUMForensicsYogeshKhatri.pdf was essential in the creation of this tool.",
       "By modifying the template file you have control of what ends up in the analyzed results.  Try creating an alternate template and passing it with the --XLSX_TEMPLATE option.",
       "This program was written by Twitter:@markbaggett and @donaldjwilliam5 because @ovie said so.",
       ])

def lookup_luid(luidval):
    LUID_interface_types = {'133': 'IF_TYPE_CES', '132': 'IF_TYPE_COFFEE', '131': 'IF_TYPE_TUNNEL', '130': 'IF_TYPE_A12MPPSWITCH', '137': 'IF_TYPE_L3_IPXVLAN', '136': 'IF_TYPE_L3_IPVLAN', '135': 'IF_TYPE_L2_VLAN', '134': 'IF_TYPE_ATM_SUBINTERFACE', '139': 'IF_TYPE_MEDIAMAILOVERIP', '138': 'IF_TYPE_DIGITALPOWERLINE', '24': 'IF_TYPE_SOFTWARE_LOOPBACK', '25': 'IF_TYPE_EON', '26': 'IF_TYPE_ETHERNET_3MBIT', '27': 'IF_TYPE_NSIP', '20': 'IF_TYPE_BASIC_ISDN', '21': 'IF_TYPE_PRIMARY_ISDN', '22': 'IF_TYPE_PROP_POINT2POINT_SERIAL', '23': 'IF_TYPE_PPP', '28': 'IF_TYPE_SLIP', '29': 'IF_TYPE_ULTRA', '4': 'IF_TYPE_DDN_X25', '8': 'IF_TYPE_ISO88024_TOKENBUS', '119': 'IF_TYPE_LAP_F', '120': 'IF_TYPE_V37', '121': 'IF_TYPE_X25_MLP', '122': 'IF_TYPE_X25_HUNTGROUP', '123': 'IF_TYPE_TRANSPHDLC', '124': 'IF_TYPE_INTERLEAVE', '125': 'IF_TYPE_FAST', '126': 'IF_TYPE_IP', '127': 'IF_TYPE_DOCSCABLE_MACLAYER', '128': 'IF_TYPE_DOCSCABLE_DOWNSTREAM', '129': 'IF_TYPE_DOCSCABLE_UPSTREAM', '118': 'IF_TYPE_HDLC', '59': 'IF_TYPE_AFLANE_8023', '58': 'IF_TYPE_FRAMERELAY_INTERCONNECT', '55': 'IF_TYPE_IEEE80212', '54': 'IF_TYPE_PROP_MULTIPLEXOR', '57': 'IF_TYPE_HIPPIINTERFACE', '56': 'IF_TYPE_FIBRECHANNEL', '51': 'IF_TYPE_SONET_VT', '50': 'IF_TYPE_SONET_PATH', '53': 'IF_TYPE_PROP_VIRTUAL', '52': 'IF_TYPE_SMDS_ICIP', '115': 'IF_TYPE_ISO88025_FIBER', '114': 'IF_TYPE_IPOVER_ATM', '88': 'IF_TYPE_ARAP', '89': 'IF_TYPE_PROP_CNLS', '111': 'IF_TYPE_STACKTOSTACK', '110': 'IF_TYPE_IPOVER_CLAW', '113': 'IF_TYPE_MPC', '112': 'IF_TYPE_VIRTUALIPADDRESS', '82': 'IF_TYPE_DS0_BUNDLE', '83': 'IF_TYPE_BSC', '80': 'IF_TYPE_ATM_LOGICAL', '81': 'IF_TYPE_DS0', '86': 'IF_TYPE_ISO88025R_DTR', '87': 'IF_TYPE_EPLRS', '84': 'IF_TYPE_ASYNC', '85': 'IF_TYPE_CNR', '3': 'IF_TYPE_HDH_1822', '7': 'IF_TYPE_IS088023_CSMACD', '108': 'IF_TYPE_PPPMULTILINKBUNDLE', '109': 'IF_TYPE_IPOVER_CDLC', '102': 'IF_TYPE_VOICE_FXS', '103': 'IF_TYPE_VOICE_ENCAP', '100': 'IF_TYPE_VOICE_EM', '101': 'IF_TYPE_VOICE_FXO', '106': 'IF_TYPE_ATM_FUNI', '107': 'IF_TYPE_ATM_IMA', '104': 'IF_TYPE_VOICE_OVERIP', '105': 'IF_TYPE_ATM_DXI', '39': 'IF_TYPE_SONET', '38': 'IF_TYPE_MIO_X25', '33': 'IF_TYPE_RS232', '32': 'IF_TYPE_FRAMERELAY', '31': 'IF_TYPE_SIP', '30': 'IF_TYPE_DS3', '37': 'IF_TYPE_ATM', '36': 'IF_TYPE_ARCNET_PLUS', '35': 'IF_TYPE_ARCNET', '34': 'IF_TYPE_PARA', '60': 'IF_TYPE_AFLANE_8025', '61': 'IF_TYPE_CCTEMUL', '62': 'IF_TYPE_FASTETHER', '63': 'IF_TYPE_ISDN', '64': 'IF_TYPE_V11', '65': 'IF_TYPE_V36', '66': 'IF_TYPE_G703_64K', '67': 'IF_TYPE_G703_2MB', '68': 'IF_TYPE_QLLC', '69': 'IF_TYPE_FASTETHER_FX', '2': 'IF_TYPE_REGULAR_1822', '6': 'IF_TYPE_ETHERNET_CSMACD', '99': 'IF_TYPE_MYRINET', '98': 'IF_TYPE_ISO88025_CRFPRINT', '91': 'IF_TYPE_TERMPAD', '90': 'IF_TYPE_HOSTPAD', '93': 'IF_TYPE_X213', '92': 'IF_TYPE_FRAMERELAY_MPI', '95': 'IF_TYPE_RADSL', '94': 'IF_TYPE_ADSL', '97': 'IF_TYPE_VDSL', '96': 'IF_TYPE_SDSL', '11': 'IF_TYPE_STARLAN', '10': 'IF_TYPE_ISO88026_MAN', '13': 'IF_TYPE_PROTEON_80MBIT', '12': 'IF_TYPE_PROTEON_10MBIT', '15': 'IF_TYPE_FDDI', '14': 'IF_TYPE_HYPERCHANNEL', '17': 'IF_TYPE_SDLC', '16': 'IF_TYPE_LAP_B', '19': 'IF_TYPE_E1', '18': 'IF_TYPE_DS1', '117': 'IF_TYPE_GIGABITETHERNET', '116': 'IF_TYPE_TDLC', '48': 'IF_TYPE_MODEM', '49': 'IF_TYPE_AAL5', '46': 'IF_TYPE_HSSI', '47': 'IF_TYPE_HIPPI', '44': 'IF_TYPE_FRAMERELAY_SERVICE', '45': 'IF_TYPE_V35', '42': 'IF_TYPE_LOCALTALK', '43': 'IF_TYPE_SMDS_DXI', '40': 'IF_TYPE_X25_PLE', '41': 'IF_TYPE_ISO88022_LLC', '1': 'IF_TYPE_OTHER', '5': 'IF_TYPE_RFC877_X25', '9': 'IF_TYPE_ISO88025_TOKENRING', '144': 'IF_TYPE_IEEE1394', '145': 'IF_TYPE_RECEIVE_ONLY', '142': 'IF_TYPE_IPFORWARD', '143': 'IF_TYPE_MSDSL', '140': 'IF_TYPE_DTM', '141': 'IF_TYPE_DCN', '77': 'IF_TYPE_LAP_D', '76': 'IF_TYPE_ISDN_U', '75': 'IF_TYPE_ISDN_S', '74': 'IF_TYPE_DLSW', '73': 'IF_TYPE_ESCON', '72': 'IF_TYPE_IBM370PARCHAN', '71': 'IF_TYPE_IEEE80211', '70': 'IF_TYPE_CHANNEL', '79': 'IF_TYPE_RSRB', '78': 'IF_TYPE_IPSWITCH'}
    inttype = struct.unpack(">H6B",format(luidval, '016x').decode("hex"))[0]
    return LUID_interface_types.get(str(inttype),'Unknown Interface type')

parser = argparse.ArgumentParser(description="Given an SRUM database it will create an XLS spreadsheet with analysis of the data in the database.")
parser.add_argument("--SRUM_INFILE","-i", help ="Specify the ESE (.dat) file to analyze. Provide a valid path to the file.")
parser.add_argument("--XLSX_TEMPLATE" ,"-t", help = "The Excel Template that specifies what data to extract from the srum database. You can create templates with ese_template.py.")
parser.add_argument("--REG_HIVE", "-r", dest="reghive", help = "If a registry hive is provided then the names of the network profiles will be resolved.")
parser.add_argument("--OUT_PATH", "-o", dest="outdir", help = "The directory to which you want to write the CSVs containing the data.")

parser.add_argument("--quiet", "-q", help = "Supress unneeded output messages.",action="store_true")

options = parser.parse_args()

interactive_mode = False
if not options.SRUM_INFILE:
    interactive_mode = True
    options.SRUM_INFILE = raw_input(r"What is the path to the SRUDB.DAT file? (Ex: \image-mount-point\Windows\system32\sru\srudb.dat) : ")
    options.XLSX_TEMPLATE = raw_input("What XLS Template should I use? (Press enter for the default SRUM_TEMPLATE.XLSX) : ")
    options.OUT_PATH = raw_input("What directory would you would like to write the CSV files to? (Press enter for the current directory) : ")
    options.reghive = raw_input("What is the full path of the SOFTWARE registry hive? Usually \image-mount-point\Windows\System32\config\SOFTWARE (or press enter to skip Network resolution) : ")

if not options.XLSX_TEMPLATE:
    options.XLSX_TEMPLATE = "SRUM_TEMPLATE.xlsx"

options.SRUM_INFILE = os.path.abspath(options.SRUM_INFILE)
if not os.path.exists(options.SRUM_INFILE):
    print "ESE File Not found: "+options.SRUM_INFILE
    abort(1)

options.OUT_PATH = os.path.abspath(options.OUT_PATH)
if not os.path.exists(options.OUT_PATH):
    print "The output directory doesn't exists: "+options.OUT_PATH
    abort(1)


options.XLSX_TEMPLATE = os.path.abspath(options.XLSX_TEMPLATE)
if not os.path.exists(options.XLSX_TEMPLATE):
    print "Template File Not found: "+options.XLSX_TEMPLATE
    abort(1)

if options.reghive:
    options.reghive = os.path.exists(options.reghive)
    if not os.path.exists(options.reghive):
        print "Registry File Not found: "+options.reghive
        abort(1)

if options.reghive:
    interface_table = load_interfaces(options.reghive)

try:
    warnings.simplefilter("ignore")
    ese_db = pyesedb.open(options.SRUM_INFILE)
except Exception as e:
    print "I could not open the specified SRUM file. Check your path and file name."
    print "Error : ", str(e)
    abort(1)

try:
    template_wb = openpyxl.load_workbook(filename=options.XLSX_TEMPLATE, read_only=True)
except Exception as e:
    print "I could not open the specified template file %s. Check your path and file name." % (options.XLSX_TEMPLATE)
    print "Error : ", str(e)
    abort(1)

id_table = load_lookups(ese_db)

sheets = template_wb.get_sheet_names()
for each_sheet in sheets:
    #open the first sheet in the template
    template_sheet = template_wb.get_sheet_by_name(each_sheet)
    #retieve the name of the ESE table to populate the sheet with from A1
    ese_template_table = template_sheet.cell("A1").value

    #retrieve the names of the ESE table columns and cell styles from row 2 and format commands from row 3 
    ese_template_fields = []
    ese_template_formats = []
    ese_template_styles = []
    #Read the first Row B & C in the template into lists so we know what data we are to extract
    for eachcolumn in range(1,template_sheet.max_column+1):
        field_name = template_sheet.cell(row = 2, column = eachcolumn).value
        if field_name == None:
            break
        field_style = copy(template_sheet.cell(row = 2, column = eachcolumn).style_array)
        format_cmd = template_sheet.cell(row = 3, column = eachcolumn).value
        ese_template_formats.append(format_cmd)
        ese_template_styles.append(field_style)
        ese_template_fields.append(field_name.strip())
    #Now open the specified table in the ESE database for this sheet
    ese_table = ese_db.get_table_by_name(ese_template_table)
    #If the table is not found it returns None
    if not ese_table:
        print "Unable to find table",each_sheet, ese_template_table
        continue

    #Now create the worksheet in the new xls file with the same name as the template
    print "\nCreating CSV for "+each_sheet

    if not options.quiet:
        try:
            ad = ads.next()
        except:
            ad = "Thanks for using srum_dump!"
    print "While you wait, did you know ...\n"+ad+"\n"

 
    csvfile = open("srum{0}.csv".format(each_sheet.replace(" ","")), "wb")
    csv_writer = csv.writer(csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)

    #Now copy the header values and header formats from the template to the new worksheet
    header_row = []
    for eachcolumn in range(1,len(ese_template_fields)+1):
        new_cell = template_sheet.cell(row = 4, column = eachcolumn ).value
        header_row.append(new_cell)

    csv_writer.writerow(header_row)

    row_num = 1 #Init to 1, first row will be 2 in spreadsheet (1 is headers)
    for ese_row in ese_getnextrow(ese_table):
        xls_row = []
        row_num += 1
        for eachcolumn,eachformat,eachstyle in zip(ese_template_fields,ese_template_formats, ese_template_styles):
            val = ese_row.get(eachcolumn,"UNABLETORETRIEVECOLUMN")
            if val=="UNABLETORETRIEVECOLUMN":
                val = "WARNING: Invalid Column Name " + eachcolumn+ " - Try one of these:"+str(ese_template_fields) + str(eachcolumn in ese_template_fields)
            val = format_value(val, eachformat)
            xls_row.append(val)
        csv_writer.writerow(xls_row)
    print "Finished processing {0}.  Writing Output File.".format(each_sheet)
    csvfile.flush()
    csvfile.close()

ese_db.close()

print "Finished!"
if interactive_mode:
    raw_input("Press enter to exit")
