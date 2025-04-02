from pathlib import Path
import openpyxl
import logging # Added for logging

from openpyxl.styles import Font
from openpyxl.cell import WriteOnlyCell, Cell
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment


# --- Logger Setup ---
logger = logging.getLogger(f"srum_dump.{__name__}")
# --- End Logger Setup ---

# Define valid number formats for cells
VALID_NUMBER_FORMATS = {
    "General": "General",
    "Text": "@",
    "Number": "#,##0.00",
    "Integer": "#,##0",
    "Percentage": "0.00%",
    "Scientific": "0.00E+00",
    "Currency": '"$"#,##0.00',
    "Euro": '"€"#,##0.00',
    "Short Date": "MM/DD/YYYY",
    "Long Date": "DD-MMM-YYYY",
    "Time": "HH:MM:SS",
    "DateTime": "MM/DD/YYYY HH:MM"
}

# Define valid font colors (Excel requires aRGB hex values)
VALID_FONT_COLORS = {
    "BLACK": "FF000000",
    "WHITE": "FFFFFFFF",
    "RED": "FFFF0000",
    "GREEN": "FF00FF00",
    "BLUE": "FF0000FF",
    "YELLOW": "FFFFFF00",
    "CYAN": "FF00FFFF",
    "MAGENTA": "FFFF00FF",
    "GRAY": "FF808080",
    "DARKRED": "FF8B0000",
    "DARKGREEN": "FF006400",
    "DARKBLUE": "FF00008B",
    "DARKYELLOW": "FF9B870C",
    "DARKCYAN": "FF008B8B",
    "DARKMAGENTA": "FF8B008B",
    "DARKGRAY": "FFA9A9A9",
    "LIGHTGRAY": "FFD3D3D3"
}

class OutputXLSX:
    """
    A class for writing Excel workbooks using openpyxl in write-only mode.
    Each worksheet is managed by a context manager.
    """
    def __init__(self):
        logger.debug("Initializing OutputXLSX.")
        self.wb = None
        self.path = None
        logger.debug("OutputXLSX initialized.")

    def new_workbook(self, path: Path):
        """
        Creates a new write-only Excel workbook.
        The caller is responsible for calling workbook.save(path) after
        writing all worksheets.

        :param path: A pathlib.Path where the workbook will eventually be saved.
        :return: An openpyxl Workbook (in write_only mode).
        """
        logger.debug(f"Called new_workbook with path: {path}")
        try:
            # Create workbook
            wb = openpyxl.Workbook() # Note: write_only=True is for optimized writing, but restricts reading/modification later. Standard mode used here.
            self.path = path.with_suffix(".xlsx")
            self.wb = wb
            logger.info(f"Created new workbook. Path set to: {self.path}")
            return wb
        except Exception as e:
            logger.exception(f"Error creating new workbook for path {path}: {e}")
            raise # Re-raise the exception

    def save(self):
        """Saves the current workbook to the path specified in new_workbook."""
        logger.debug(f"Called save for workbook path: {self.path}")
        # Remove the default "Sheet" if it exists 
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        if not self.wb or not self.path:
            err_msg = "Workbook or path not initialized. Call new_workbook first."
            logger.error(err_msg)
            raise Exception(err_msg)
        try:
            logger.info(f"Saving workbook to: {self.path}")
            self.wb.save(self.path)
            logger.info("Workbook saved successfully.")
        except Exception as e:
            logger.exception(f"Error saving workbook to {self.path}: {e}")
            raise # Re-raise the exception


    class XLSXWorksheetContext:
        """Context manager for handling individual worksheets."""
        def __init__(self, workbook, worksheet_name: str, column_headers: list):
            # Use a nested logger name
            self.logger = logging.getLogger(f"srum_dump.{__name__}.XLSXWorksheetContext")
            self.logger.debug(f"Initializing XLSXWorksheetContext for sheet: '{worksheet_name}'")
            try:
                self.workbook = workbook
                self.worksheet_name = worksheet_name
                self.column_headers = column_headers
                self.worksheet = None
                self.logger.debug(f"Context initialized for sheet '{worksheet_name}' with {len(column_headers)} headers.")
            except Exception as e:
                 self.logger.exception(f"Error during XLSXWorksheetContext initialization: {e}")
                 raise

        def __enter__(self):
            """Creates the worksheet and sets up headers/formatting upon entering the context."""
            self.logger.debug(f"Entering context for sheet: '{self.worksheet_name}'")
            try:
                # Create a new worksheet and write the header row.
                self.logger.info(f"Creating sheet: '{self.worksheet_name}'")
                self.worksheet = self.workbook.create_sheet(title=self.worksheet_name)
                self.logger.debug(f"Appending headers: {self.column_headers}")
                self.worksheet.append(self.column_headers)

                # Bold first row
                # self.logger.debug("Bolding header row.")
                # for col_idx, _ in enumerate(self.column_headers, start=1):
                #     cell_ref = f"{get_column_letter(col_idx)}1"
                #     cell = self.worksheet[cell_ref]
                #     cell.font = Font(bold=True)

                # Bold, center, and add borders to header row
                thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                    top=Side(style="thin"), bottom=Side(style="thin"))
                self.logger.debug("Formatting header row.")

                for col_idx, _ in enumerate(self.column_headers, start=1):
                    cell_ref = f"{get_column_letter(col_idx)}1"
                    cell = self.worksheet[cell_ref]
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = thin_border

                # Freeze first row
                self.worksheet.freeze_panes = "A2"
                self.logger.debug("Froze header row (A2).")

                # Set AutoFilter for all columns (only header row)
                if self.column_headers: # Only set filter if there are headers
                    last_col = get_column_letter(len(self.column_headers))
                    filter_ref = f"A1:{last_col}1"
                    self.worksheet.auto_filter.ref = filter_ref
                    self.logger.debug(f"Set autofilter reference to: {filter_ref}")
                else:
                    self.logger.warning("No column headers provided, skipping autofilter setup.")

                self.logger.debug(f"Worksheet '{self.worksheet_name}' setup complete.")
                return self.worksheet
            except Exception as e:
                 self.logger.exception(f"Error during __enter__ for sheet '{self.worksheet_name}': {e}")
                 # Ensure worksheet is None if setup fails to prevent issues in __exit__
                 self.worksheet = None
                 raise # Re-raise the exception

        def __exit__(self, exc_type, exc_val, exc_tb):
            """Cleans up worksheet reference upon exiting the context."""
            self.logger.debug(f"Adjusting column widths for sheet: '{self.worksheet_name}'")
            try:
                # Adjust column widths after data is populated
                self.adjust_column_widths(max_size=100, sample_rows=10)
                
                # Save the workbook or other cleanup (if applicable)
                self.logger.info(f"Completed operations on sheet: '{self.worksheet_name}'")
            except Exception as e:
                self.logger.exception(f"Error during exit: {str(e)}")
                raise Exception(e)

            if exc_type:
                # Log any exception that occurred within the 'with' block
                self.logger.error(f"Exception occurred within worksheet context '{self.worksheet_name}': {exc_type.__name__}: {exc_val}", exc_info=(exc_type, exc_val, exc_tb))

            # No explicit close needed for openpyxl worksheets in standard mode.
            # Dereferencing helps garbage collection if needed, but not strictly required.
            self.worksheet = None
            self.logger.debug(f"Worksheet reference cleared for '{self.worksheet_name}'.")
            # Return False to propagate exceptions, True to suppress (default is False/None)

        def adjust_column_widths(self, max_size=100, sample_rows=10):
            """Adjusts column widths based on the longest value in the first `sample_rows` rows."""
            self.logger.debug(f"Adjusting column widths for sheet: '{self.worksheet_name}'")
            if not self.column_headers:
                self.logger.warning("No column headers present, skipping width adjustment.")
                return

            for col_idx, header in enumerate(self.column_headers, start=1):
                column_letter = get_column_letter(col_idx)
                max_length = len(str(header))  # Start with header length

                # Sample the first `sample_rows` rows (including header)
                for row in self.worksheet.iter_rows(min_row=1, max_row=min(sample_rows, self.worksheet.max_row),
                                                    min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        try:
                            cell_value = str(cell.value)
                            if len(cell_value) > max_length:
                                max_length = len(cell_value)
                        except (TypeError, AttributeError):
                            continue
                
                # Set column width with padding, capped at max_size
                adjusted_width = min(max_length + 2, max_size)
                self.worksheet.column_dimensions[column_letter].width = adjusted_width
                self.logger.debug(f"Set width of column {column_letter} to {adjusted_width}")

    def new_worksheet(self, workbook, worksheet_name: str, column_headers: list):
        """
        Returns a context manager for a new worksheet in the given workbook.
        Usage:
            with instance.new_worksheet(workbook, "Sheet1", headers) as ws:
                instance.new_entry(ws, row)
                ...
        :param workbook: The openpyxl Workbook object.
        :param worksheet_name: Name for the worksheet.
        :param column_headers: List of column header strings.
        :return: A context manager for the worksheet.
        """
        logger.debug(f"Creating new worksheet context for sheet: '{worksheet_name}'")
        # Validate inputs briefly
        if not isinstance(workbook, openpyxl.workbook.workbook.Workbook):
             err_msg = "Invalid workbook object provided to new_worksheet."
             logger.error(err_msg)
             raise TypeError(err_msg)
        if not isinstance(worksheet_name, str) or not worksheet_name:
             err_msg = "Worksheet name must be a non-empty string."
             logger.error(err_msg)
             raise ValueError(err_msg)
        if not isinstance(column_headers, list):
             err_msg = "Column headers must be a list."
             logger.error(err_msg)
             raise TypeError(err_msg)

        return self.XLSXWorksheetContext(workbook, worksheet_name, column_headers)


    def new_entry(self, worksheet, entry: list, format_options: list = None): # Default to None
        """
        Appends a row of data to the given worksheet with formatting.

        :param worksheet: The worksheet object (as yielded by the context manager).
        :param entry: List of cell values.
        :param format_options: List of tuples (cell_format, font_format), same length as entry, or None.
        """
        # Use an empty list if format_options is None
        format_options = format_options or []
        logger.debug(f"Called new_entry for sheet '{worksheet.title}' with {len(entry)} values. Format options provided: {bool(format_options)}")
        try:
            # Write data to worksheet
            worksheet.append(entry)
            logger.debug(f"Appended entry to sheet '{worksheet.title}'.")

            # If no formatting specified or list is empty, return early
            if not format_options or not any(format_options):
                logger.debug("No format options provided or all are None, skipping formatting.")
                return

            # Get the row index of the newly added row
            row_idx = worksheet.max_row
            logger.debug(f"Applying formatting to row {row_idx}.")

            # Apply formatting cell by cell for the new row
            for col_idx_zero_based, format_value in enumerate(format_options):
                col_idx_one_based = col_idx_zero_based + 1

                if not format_value:  # Skip if no format specified for this cell
                    # logger.debug(f"No format for row {row_idx}, col {col_idx_one_based}.") # Can be verbose
                    continue

                # Ensure format_value is a tuple/list of length 2
                if not isinstance(format_value, (list, tuple)) or len(format_value) != 2:
                     logger.warning(f"Invalid format_value structure at row {row_idx}, col {col_idx_one_based}: {format_value}. Expected (cell_format, font_format). Skipping.")
                     continue

                cell_format, font_format = format_value
                logger.debug(f"Processing format for row {row_idx}, col {col_idx_one_based}: cell='{cell_format}', font='{font_format}'")

                try:
                    # Get the cell reference (safer access)
                    cell = worksheet.cell(row=row_idx, column=col_idx_one_based)
                    cell.alignment = Alignment(wrap_text=True)

                    # Validate and apply cell number format
                    if cell_format:
                        if cell_format in VALID_NUMBER_FORMATS:
                            cell.number_format = VALID_NUMBER_FORMATS[cell_format]
                            logger.debug(f"Applied number format '{VALID_NUMBER_FORMATS[cell_format]}' to cell {cell.coordinate}")
                        else:
                            logger.error(f"Invalid cell format specified: '{cell_format}' for cell {cell.coordinate}. Skipping number format.")
                            # Optionally raise an error: raise ValueError(f"Invalid cell format: {cell_format}")

                    # Apply font formatting
                    if font_format:
                        font_parts = font_format.split(":")
                        font_style = font_parts[0].upper() # Ensure uppercase for comparison
                        font_color_name = font_parts[1].upper() if len(font_parts) > 1 else "BLACK"

                        # Validate font color name
                        if font_color_name in VALID_FONT_COLORS:
                            font_color_hex = VALID_FONT_COLORS[font_color_name]
                            is_bold = (font_style == "BOLD")
                            # Apply the font style and color
                            cell.font = Font(bold=is_bold, color=font_color_hex)
                            logger.debug(f"Applied font (Bold={is_bold}, Color={font_color_name}/{font_color_hex}) to cell {cell.coordinate}")
                        else:
                            logger.error(f"Invalid font color specified: '{font_color_name}' for cell {cell.coordinate}. Skipping font format.")
                            # Optionally raise an error: raise ValueError(f"Invalid font color: {font_color}")

                except Exception as cell_format_ex:
                    # Log error applying format to a specific cell but continue with others
                    logger.exception(f"Error applying format to cell at row {row_idx}, col {col_idx_one_based}: {cell_format_ex}")

        except Exception as e:
            logger.exception(f"Error in new_entry for sheet '{worksheet.title}': {e}")
            # Decide if the error should halt processing or just be logged
            # raise # Option: re-raise the exception
