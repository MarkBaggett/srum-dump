import openpyxl
import struct
import codecs
import re
from openpyxl.styles import Font
from openpyxl.cell import WriteOnlyCell, Cell

from srum_dump.utils import file_timestamp

class XLSXHandler:
    def __init__(self, template_file, output_file):
        self.template_file = template_file
        self.output_file = output_file
        self.template_workbook = None
        self.output_workbook = openpyxl.Workbook()
        self.template_tables = {}
        self.template_lookups = {}

    def load_template(self):
        self.template_workbook = openpyxl.load_workbook(filename=self.template_file)
        self.template_tables = self.load_template_tables(self.template_workbook)
        self.template_lookups = self.load_template_lookups(self.template_workbook)

    def load_template_tables(self, template_workbook):
        template_tables = {}
        for each_sheet in template_workbook.get_sheet_names():
            if each_sheet.lower().startswith("lookup-"):
                lookupname = each_sheet.split("-")[1]
                template_sheet = template_workbook.get_sheet_by_name(each_sheet)
                lookup_table = {}
                for eachrow in range(1, template_sheet.max_row + 1):
                    value = template_sheet.cell(row=eachrow, column=1).value
                    description = template_sheet.cell(row=eachrow, column=2).value
                    lookup_table[value] = description
                self.template_lookups[lookupname] = lookup_table
            else:
                template_sheet = template_workbook.get_sheet_by_name(each_sheet)
                ese_template_table = template_sheet.cell(row=1, column=1).value
                template_field = {}
                for eachcolumn in range(1, template_sheet.max_column + 1):
                    field_name = template_sheet.cell(row=2, column=eachcolumn).value
                    if field_name is None:
                        break
                    template_style = template_sheet.cell(row=4, column=eachcolumn).style
                    template_format = template_sheet.cell(row=3, column=eachcolumn).value
                    template_value = template_sheet.cell(row=4, column=eachcolumn).value
                    if not template_value:
                        template_value = field_name
                    template_field[field_name] = (template_style, template_format, template_value)
                template_tables[ese_template_table] = (each_sheet, template_field)
        return template_tables

    def load_template_lookups(self, template_workbook):
        template_lookups = {}
        for each_sheet in template_workbook.get_sheet_names():
            if each_sheet.lower().startswith("lookup-"):
                lookupname = each_sheet.split("-")[1]
                template_sheet = template_workbook.get_sheet_by_name(each_sheet)
                lookup_table = {}
                for eachrow in range(1, template_sheet.max_row + 1):
                    value = template_sheet.cell(row=eachrow, column=1).value
                    description = template_sheet.cell(row=eachrow, column=2).value
                    lookup_table[value] = description
                template_lookups[lookupname] = lookup_table
        return template_lookups

    def format_output(self, val, eachformat, eachstyle, xls_sheet):
        """Returns a excel cell with the data formated as specified in the template table"""
        new_cell = WriteOnlyCell(xls_sheet, value = "init")
        new_cell.style = eachstyle
        if val==None:
            val="None"
        elif eachformat in [None, "OLE"]:
            pass
        elif eachformat.startswith("OLE:"):
            val = val.strftime(eachformat[4:])
        elif eachformat=="FILE":
            val = file_timestamp(val)
            new_cell.number_format = 'YYYY MMM DD'
        elif eachformat.startswith("FILE:"):
            val = file_timestamp(val)
            val = val.strftime(eachformat[5:])
        elif eachformat.lower().startswith("lookup-"):
            lookup_name = eachformat.split("-")[1]
            if lookup_name in self.template_lookups:
                lookup_table = self.template_lookups.get(lookup_name,{})
                val = lookup_table.get(val,val)
        elif eachformat.lower() == "lookup_id":
            val = id_table.get(val, "No match in srum lookup table for %s" % (val))
        elif eachformat.lower() == "lookup_luid":
            inttype = struct.unpack(">H6B", codecs.decode(format(val,'016x'),'hex'))[0]
            val = self.template_lookups.get("LUID Interfaces",{}).get(inttype,"")
        elif eachformat.lower() == "seconds":
            val = val/86400.0
            new_cell.number_format = 'dd hh:mm:ss'
        elif eachformat.lower() == "md5":
            val = hashlib.md5(str(val)).hexdigest()
        elif eachformat.lower() == "sha1":
            val = hashlib.sha1(str(val)).hexdigest()
        elif eachformat.lower() == "sha256":
            val = hashlib.sha256(str(val)).hexdigest()
        elif eachformat.lower() == "base16":
            if type(val)==int:
                val = hex(val)
            else:
                val = format(val,"08x")
        elif eachformat.lower() == "base2":
            if type(val)==int:
                val = format(val,"032b")
            else:
                try:
                    val = int(str(val),2)
                except :
                    val = val
        elif eachformat.lower() == "interface_id" and options.reghive:
            val = interface_table.get(str(val),"")
        elif eachformat.lower() == "interface_id" and not options.reghive:
            val = val
        else:
            val = val
        try:
            new_cell.value = val
        except:
            new_cell.value = re.sub(r'[\000-\010]|[\013-\014]|[\016-\037]|[\x00-\x1f\x7f-\x9f]|[\uffff]',"",val)
        return new_cell


    def save_output(self):
        self.output_workbook.save(self.output_file)
